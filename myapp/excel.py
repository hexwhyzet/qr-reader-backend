from datetime import datetime

import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Font, PatternFill, Alignment, colors

from myapp.models import Point, Guard, Round


def adjust_col_width(worksheet):
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column].width = adjusted_width


def style_header(cell):
    header_fill = PatternFill(start_color="0c4b33", end_color="0c4b33", fill_type="solid")
    header_font = Font(bold=False, color=colors.WHITE)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")


def style_list(cell):
    if not cell.row % 2:
        data_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        cell.fill = data_fill
    cell.font = Font(color="000000")
    cell.alignment = Alignment(horizontal="left", vertical="center")


def style_separator(cell):
    data_fill = PatternFill(start_color="A3DDC7", end_color="A3DDC7", fill_type="solid")
    cell.fill = data_fill


def fire_extinguishers(*args):
    fire_extinguishers = Point.objects.filter(point_type='fire_extinguisher')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Информация об огнетушителях"

    headers = ["id", "Название точки", "Срок годности"]
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        style_header(ws.cell(row=1, column=col))

    for point in fire_extinguishers:
        expiration_date = point.expiration_date
        row = [point.id, point.name, expiration_date.strftime('%Y-%m-%d') if expiration_date else "N/A"]

        ws.append(row)

        for col_num, value in enumerate(row, start=1):
            style_list(ws.cell(row=ws.max_row, column=col_num))

        if expiration_date and expiration_date <= datetime.now().date():
            cell = ws.cell(row=ws.max_row, column=3)
            cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
            cell.font = Font(color="FFFFFF")

    adjust_col_width(ws)

    current_date = datetime.now().strftime('%Y-%m-%d')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="fire_extinguishers_{current_date}.xlsx"'

    wb.save(response)
    return response


def guards_stats(*args):
    wb = openpyxl.Workbook()
    ws_first = True

    guards = Guard.objects.all()
    for guard in guards:
        if ws_first:
            ws = wb.active
            ws.title = guard.name
            ws_first = False
        else:
            ws = wb.create_sheet(title=guard.name)

        headers = ["id обхода", "Точка обхода", "Время обхода"]
        ws.append(headers)

        for col_num in range(1, len(headers) + 1):
            style_header(ws.cell(row=1, column=col_num))

        rounds = Round.objects.filter(guard=guard).order_by('-created_at').all()
        for round_obj in rounds:
            visits = round_obj.visits.order_by('-created_at').all()
            if visits.exists():
                for visit in visits:
                    row = [
                        round_obj.id,
                        visit.point.name,
                        visit.created_at.strftime('%Y-%m-%d %H:%M:%S')
                    ]

                    ws.append(row)
                    for col_num, value in enumerate(row, start=1):
                        style_list(ws.cell(row=ws.max_row, column=col_num))

            ws.append([" "])
            for col in range(1, ws.max_column + 1):
                style_separator(ws.cell(row=ws.max_row, column=col))

        adjust_col_width(ws)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )

    current_date = datetime.now().strftime('%Y-%m-%d')

    response['Content-Disposition'] = f'attachment; filename="rounds_export_{current_date}.xlsx"'

    wb.save(response)
    return response
